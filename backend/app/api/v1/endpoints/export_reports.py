"""
backend/app/api/v1/endpoints/export_reports.py

Thêm vào api.py:
    from app.api.v1.endpoints import export_reports
    router.include_router(export_reports.router, prefix="/statistics", tags=["export"])

Tạo thư mục fonts/ cạnh file này:
    fonts/FreeSans.ttf
    fonts/FreeSansBold.ttf
    fonts/FreeSansOblique.ttf
"""

import os as _os
from io import BytesIO
from datetime import datetime

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.db.base import get_db
from app.models.user import User
from app.models.intern_batch import InternBatch
from app.models.intern_profile import InternProfile
from app.models.task import Task
from app.models.evaluation import Evaluation
from app.core.deps import get_current_hr

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter()

# ─── Font setup ───────────────────────────────────────────────────────────────

_FONT_DIR = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "fonts")
_FONT_REG = False

def _register_fonts():
    global _FONT_REG
    if _FONT_REG:
        return
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.pdfbase.pdfmetrics import registerFontFamily

    def fp(name):
        return _os.path.join(_FONT_DIR, name)

    pdfmetrics.registerFont(TTFont("DV",   fp("FreeSans.ttf")))
    pdfmetrics.registerFont(TTFont("DV-B", fp("FreeSansBold.ttf")))
    pdfmetrics.registerFont(TTFont("DV-I", fp("FreeSansOblique.ttf")))
    registerFontFamily("DV", normal="DV", bold="DV-B", italic="DV-I", boldItalic="DV-B")
    _FONT_REG = True


# ─── Excel helpers ────────────────────────────────────────────────────────────

MB_RED    = "CC0000"
MB_WHITE  = "FFFFFF"
MB_DARK   = "1A1A2E"
MB_GREEN  = "1E8449"
MB_YELLOW = "F9A825"

_THIN = Side(style="thin", color="CCCCCC")
_THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HDR_BORDER  = Border(
    left=Side(style="medium", color=MB_RED),   right=Side(style="medium", color=MB_RED),
    top=Side(style="medium", color=MB_RED),    bottom=Side(style="medium", color=MB_RED),
)


def _apply_table_header(ws, row: int, cols: list):
    for ci, (label, width) in enumerate(cols, 1):
        cell = ws.cell(row=row, column=ci, value=label)
        cell.font      = Font(bold=True, color=MB_WHITE, name="Arial", size=10)
        cell.fill      = PatternFill("solid", fgColor=MB_RED)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _HDR_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[row].height = 28


def _style_data_row(ws, row: int, n: int, alt=False):
    fill = PatternFill("solid", fgColor="F9F9F9" if alt else "FFFFFF")
    for c in range(1, n + 1):
        cell = ws.cell(row=row, column=c)
        cell.border    = _THIN_BORDER
        cell.fill      = fill
        cell.font      = Font(name="Arial", size=10)
        cell.alignment = Alignment(vertical="center")


# ─── GET /statistics/export-excel ─────────────────────────────────────────────

@router.get("/export-excel")
def export_excel(
    batch_id: int | None = None,
    current_user: User = Depends(get_current_hr),
    db: Session = Depends(get_db),
):
    gen_time = datetime.now().strftime("%d/%m/%Y %H:%M")

    batches_q  = db.query(InternBatch).all()
    profiles_q = db.query(InternProfile).all()
    tasks_q    = db.query(Task).all()
    evals_q    = db.query(Evaluation).filter(Evaluation.total_score.isnot(None)).all()

    if batch_id:
        bids       = {p.user_id for p in profiles_q if p.batch_id == batch_id}
        tasks_q    = [t for t in tasks_q    if t.batch_id == batch_id]
        evals_q    = [e for e in evals_q    if e.intern_id in bids]
        profiles_q = [p for p in profiles_q if p.batch_id == batch_id]
        filter_info = next((b.batch_name for b in batches_q if b.id == batch_id), "Tất cả đợt")
    else:
        filter_info = "Tất cả đợt"

    wb = openpyxl.Workbook()

    # Sheet 1: Bìa
    ws = wb.active
    ws.title = "Bìa"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 30
    ws.row_dimensions[2].height = 60
    for col in range(1, 4):
        ws.cell(row=2, column=col).fill = PatternFill("solid", fgColor=MB_RED)
    ws.merge_cells("B2:C2")
    tc = ws.cell(row=2, column=2, value="MB MILITARY BANK\nHỆ THỐNG QUẢN LÝ THỰC TẬP")
    tc.font      = Font(bold=True, color=MB_WHITE, name="Arial", size=16)
    tc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[5].height = 36
    ws.merge_cells("B5:C5")
    t = ws.cell(row=5, column=2, value="BÁO CÁO TỔNG HỢP TÌNH HÌNH THỰC TẬP SINH")
    t.font      = Font(bold=True, color=MB_DARK, name="Arial", size=14)
    t.alignment = Alignment(horizontal="center", vertical="center")
    t.fill      = PatternFill("solid", fgColor="FDECEA")
    for i, (label, value) in enumerate([
        ("Đơn vị lập:", "Phòng Nhân sự (HR) — MB Military Bank"),
        ("Ngày xuất:", gen_time),
        ("Phạm vi:", filter_info),
        ("Mã BC:", f"MB-HR-INT-{datetime.now().strftime('%Y%m%d')}-001"),
        ("Phân loại:", "NỘI BỘ — CONFIDENTIAL"),
    ], 8):
        ws.row_dimensions[i].height = 22
        lc = ws.cell(row=i, column=2, value=label)
        lc.font = Font(bold=True, name="Arial", size=10, color="555555")
        vc = ws.cell(row=i, column=3, value=value)
        vc.font = Font(name="Arial", size=10)

    # Sheet 2: Tổng quan
    total_tasks = len(tasks_q)
    approved    = sum(1 for t in tasks_q if t.status == "approved")
    comp_rate   = round(approved / total_tasks * 100, 1) if total_tasks else 0
    avg_score   = round(sum(e.total_score for e in evals_q) / len(evals_q), 1) if evals_q else 0
    ws2 = wb.create_sheet("Tổng quan")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 4
    ws2.column_dimensions["B"].width = 40
    ws2.column_dimensions["C"].width = 22
    ws2.row_dimensions[2].height = 32
    ws2.merge_cells("B2:C2")
    t2 = ws2.cell(row=2, column=2, value="I. TỔNG QUAN HỆ THỐNG")
    t2.font      = Font(bold=True, name="Arial", size=13, color=MB_WHITE)
    t2.fill      = PatternFill("solid", fgColor=MB_RED)
    t2.alignment = Alignment(horizontal="center", vertical="center")
    for i, (label, val) in enumerate([
        ("Tổng thực tập sinh", len(profiles_q)),
        ("Đã đánh giá", len(evals_q)),
        ("Tổng nhiệm vụ", total_tasks),
        ("Đã duyệt", approved),
        ("Tỷ lệ hoàn thành (%)", f"{comp_rate}%"),
        ("Điểm trung bình", f"{avg_score}/10"),
        ("Đợt đang mở", sum(1 for b in batches_q if b.status == "open")),
    ], 4):
        ws2.row_dimensions[i].height = 20
        lc = ws2.cell(row=i, column=2, value=label)
        lc.font   = Font(name="Arial", size=10)
        lc.border = _THIN_BORDER
        lc.fill   = PatternFill("solid", fgColor="F5F5F5" if i % 2 == 0 else "FFFFFF")
        vc = ws2.cell(row=i, column=3, value=val)
        vc.font      = Font(bold=True, name="Arial", size=10, color=MB_RED)
        vc.alignment = Alignment(horizontal="center")
        vc.border    = _THIN_BORDER
        vc.fill      = PatternFill("solid", fgColor="F5F5F5" if i % 2 == 0 else "FFFFFF")

    # Sheet 3: Theo đợt
    ws3 = wb.create_sheet("Theo đợt")
    ws3.sheet_view.showGridLines = False
    _apply_table_header(ws3, 1, [
        ("Tên đợt", 28), ("Trạng thái", 14), ("Ngày bắt đầu", 15),
        ("Ngày kết thúc", 15), ("Số TTS", 10), ("Tổng NV", 10),
        ("Đã duyệt", 12), ("Tỷ lệ HT (%)", 14), ("Đã ĐG", 10), ("Điểm TB", 10),
    ])
    for i, b in enumerate(batches_q, 2):
        b_iids  = [p.user_id for p in profiles_q if p.batch_id == b.id]
        b_tasks = [t for t in tasks_q if t.batch_id == b.id]
        b_evs   = [e for e in evals_q if e.intern_id in b_iids]
        bt = len(b_tasks); ba = sum(1 for t in b_tasks if t.status == "approved")
        br = round(ba / bt * 100, 1) if bt else 0
        bavg = round(sum(e.total_score for e in b_evs) / len(b_evs), 1) if b_evs else None
        _style_data_row(ws3, i, 10, i % 2 == 0)
        for j, v in enumerate([
            b.batch_name,
            "Đang mở" if b.status == "open" else "Đã đóng",
            b.start_date.strftime("%d/%m/%Y") if b.start_date else "",
            b.end_date.strftime("%d/%m/%Y") if b.end_date else "",
            len(b_iids), bt, ba, f"{br}%", len(b_evs),
            bavg if bavg is not None else "—",
        ], 1):
            c = ws3.cell(row=i, column=j, value=v)
            c.alignment = Alignment(horizontal="center" if j > 1 else "left", vertical="center")
            if j == 8:
                c.font = Font(name="Arial", size=10, bold=True,
                    color=MB_GREEN if br >= 70 else (MB_YELLOW if br >= 40 else "CC0000"))
        ws3.row_dimensions[i].height = 20

    # Sheet 4: Thực tập sinh
    ws4 = wb.create_sheet("Thực tập sinh")
    ws4.sheet_view.showGridLines = False
    _apply_table_header(ws4, 1, [
        ("Họ tên TTS", 28), ("Đợt", 22), ("Tổng NV", 10),
        ("Đã duyệt", 10), ("Đã nộp", 10), ("Quá hạn", 10), ("Tỷ lệ HT (%)", 14),
    ])
    uc: dict = {}
    def gu(uid):
        if uid not in uc:
            uc[uid] = db.query(User).filter(User.id == uid).first()
        return uc[uid]
    bmap = {b.id: b.batch_name for b in batches_q}
    for i, prof in enumerate(profiles_q, 2):
        user = gu(prof.user_id)
        if not user: continue
        pt = [t for t in tasks_q if t.intern_id == prof.user_id]
        pn = len(pt)
        pa = sum(1 for t in pt if t.status == "approved")
        ps = sum(1 for t in pt if t.status == "submitted")
        po = sum(1 for t in pt if t.status == "overdue")
        pr = round(pa / pn * 100, 1) if pn else 0
        _style_data_row(ws4, i, 7, i % 2 == 0)
        for j, v in enumerate([user.full_name, bmap.get(prof.batch_id,"—"), pn, pa, ps, po, f"{pr}%"], 1):
            c = ws4.cell(row=i, column=j, value=v)
            c.alignment = Alignment(horizontal="left" if j <= 2 else "center", vertical="center")
        ws4.row_dimensions[i].height = 20

    # Sheet 5: Đánh giá
    ws5 = wb.create_sheet("Đánh giá")
    ws5.sheet_view.showGridLines = False
    _apply_table_header(ws5, 1, [
        ("Họ tên TTS", 26), ("Thái độ", 11), ("Kỷ luật", 11),
        ("Học hỏi", 11), ("Kỹ năng", 11), ("Kết quả NV", 12),
        ("Tổng điểm", 12), ("Xếp loại", 14), ("Mentor", 22), ("Ngày ĐG", 14),
    ])
    rank_colors = {
        "Xuất sắc": "7B1FA2", "Giỏi": "1B5E20", "Khá": "0D47A1",
        "Trung bình": "F57F17", "Yếu": "B71C1C",
    }
    for i, ev in enumerate(sorted(evals_q, key=lambda x: -(x.total_score or 0)), 2):
        iu = gu(ev.intern_id); mu = gu(ev.mentor_id)
        cs = ev.criteria_scores or {}
        _style_data_row(ws5, i, 10, i % 2 == 0)
        for j, v in enumerate([
            iu.full_name if iu else f"ID:{ev.intern_id}",
            cs.get("attitude"), cs.get("discipline"), cs.get("learning"),
            cs.get("skills"), cs.get("task_result"), ev.total_score, ev.ranking,
            mu.full_name if mu else f"ID:{ev.mentor_id}",
            ev.created_at.strftime("%d/%m/%Y") if ev.created_at else "",
        ], 1):
            c = ws5.cell(row=i, column=j, value=v)
            c.alignment = Alignment(horizontal="center" if j != 1 else "left", vertical="center")
            if j == 8 and v in rank_colors:
                c.font = Font(name="Arial", size=10, bold=True, color=rank_colors[v])
            if j == 7 and v is not None:
                c.font = Font(name="Arial", size=10, bold=True,
                    color=MB_GREEN if v >= 8 else (MB_YELLOW if v >= 6.5 else "CC0000"))
        ws5.row_dimensions[i].height = 20

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"bao-cao-thuc-tap-{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"})


# ─── PDF canvas helpers ────────────────────────────────────────────────────────

def _draw_rect(c, x, y, w, h, fill_color=None, stroke_color=None, radius=0):
    if fill_color:
        c.setFillColor(fill_color)
    if stroke_color:
        c.setStrokeColor(stroke_color)
    else:
        c.setStrokeColor(fill_color or _pdf_colors()["WHITE"])
    if radius:
        c.roundRect(x, y, w, h, radius, fill=1 if fill_color else 0, stroke=1 if stroke_color else 0)
    else:
        c.rect(x, y, w, h, fill=1 if fill_color else 0, stroke=1 if stroke_color else 0)


def _pdf_colors():
    from reportlab.lib import colors
    return {
        "RED":   colors.HexColor("#CC0000"),
        "WHITE": colors.white,
        "DARK":  colors.HexColor("#1A1A2E"),
        "GREY":  colors.HexColor("#888888"),
        "LGREY": colors.HexColor("#CCCCCC"),
        "LIGHT": colors.HexColor("#F5F5F5"),
        "GREEN": colors.HexColor("#1E8449"),
    }


def _draw_table(c, x, y, rows, col_widths, row_height=18, header_color=None, font_reg="DV", font_bold="DV-B", font_size=9):
    """Draw a table using pure canvas calls — no XML parser."""
    from reportlab.lib import colors as rl_colors
    clr = _pdf_colors()
    RED   = clr["RED"]
    WHITE = clr["WHITE"]
    LIGHT = clr["LIGHT"]
    LGREY = clr["LGREY"]

    hdr_color = header_color or RED
    total_w = sum(col_widths)

    for ri, row in enumerate(rows):
        row_y = y - ri * row_height
        # Row background
        if ri == 0:
            c.setFillColor(hdr_color)
            c.rect(x, row_y - row_height, total_w, row_height, fill=1, stroke=0)
        else:
            bg = LIGHT if ri % 2 == 0 else WHITE
            c.setFillColor(bg)
            c.rect(x, row_y - row_height, total_w, row_height, fill=1, stroke=0)

        # Row border
        c.setStrokeColor(LGREY)
        c.setLineWidth(0.4)
        c.rect(x, row_y - row_height, total_w, row_height, fill=0, stroke=1)

        # Cells
        cx = x
        for ci, (cell_val, col_w) in enumerate(zip(row, col_widths)):
            # Cell vertical divider
            c.setStrokeColor(LGREY)
            c.setLineWidth(0.4)
            c.line(cx, row_y, cx, row_y - row_height)

            # Text
            text = str(cell_val) if cell_val is not None else "—"
            fn   = font_bold if ri == 0 else font_reg
            fc   = WHITE     if ri == 0 else rl_colors.black
            c.setFont(fn, font_size)
            c.setFillColor(fc)

            # Truncate if too long
            max_w = col_w - 4
            while c.stringWidth(text, fn, font_size) > max_w and len(text) > 1:
                text = text[:-1]

            # Center horizontally
            text_w = c.stringWidth(text, fn, font_size)
            text_x = cx + (col_w - text_w) / 2
            text_y = row_y - row_height + (row_height - font_size) / 2 + 1
            c.drawString(text_x, text_y, text)

            cx += col_w

        # Last vertical line
        c.setStrokeColor(LGREY)
        c.line(cx, row_y, cx, row_y - row_height)

    # Outer border
    c.setStrokeColor(RED)
    c.setLineWidth(0.8)
    total_h = len(rows) * row_height
    c.rect(x, y - total_h, total_w, total_h, fill=0, stroke=1)


def _wrap_text(c, text, font, size, max_width):
    """Split text into lines that fit max_width."""
    words  = text.split()
    lines  = []
    line   = ""
    for word in words:
        test = (line + " " + word).strip()
        if c.stringWidth(test, font, size) <= max_width:
            line = test
        else:
            if line:
                lines.append(line)
            line = word
    if line:
        lines.append(line)
    return lines


# ─── GET /statistics/export-pdf ───────────────────────────────────────────────

@router.get("/export-pdf")
def export_pdf(
    batch_id: int | None = None,
    current_user: User = Depends(get_current_hr),
    db: Session = Depends(get_db),
):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.pdfgen import canvas as rl_canvas

    _register_fonts()
    clr = _pdf_colors()
    RED   = clr["RED"]
    WHITE = clr["WHITE"]
    DARK  = clr["DARK"]
    GREY  = clr["GREY"]
    LGREY = clr["LGREY"]
    LIGHT = clr["LIGHT"]

    gen_time = datetime.now().strftime("%d/%m/%Y %H:%M")
    ref_no   = f"MB-HR-INT-{datetime.now().strftime('%Y%m%d')}-001"
    W, H     = A4
    ML = MR  = 2 * cm
    MT = 2 * cm
    content_w = W - ML - MR

    # ── Data ──
    batches_q  = db.query(InternBatch).all()
    profiles_q = db.query(InternProfile).all()
    tasks_q    = db.query(Task).all()
    evals_q    = db.query(Evaluation).filter(Evaluation.total_score.isnot(None)).all()

    if batch_id:
        bids       = {p.user_id for p in profiles_q if p.batch_id == batch_id}
        tasks_f    = [t for t in tasks_q    if t.batch_id == batch_id]
        evals_f    = [e for e in evals_q    if e.intern_id in bids]
        profs_f    = [p for p in profiles_q if p.batch_id == batch_id]
        filter_info = next((b.batch_name for b in batches_q if b.id == batch_id), "Tất cả đợt")
    else:
        tasks_f = tasks_q; evals_f = evals_q; profs_f = profiles_q
        filter_info = "Tất cả đợt"

    total_tasks = len(tasks_f)
    approved    = sum(1 for t in tasks_f if t.status == "approved")
    comp_rate   = round(approved / total_tasks * 100, 1) if total_tasks else 0
    avg_score   = round(sum(e.total_score for e in evals_f) / len(evals_f), 1) if evals_f else 0
    overdue_cnt = sum(1 for t in tasks_f if t.status == "overdue")

    uc: dict = {}
    def gu(uid):
        if uid not in uc:
            uc[uid] = db.query(User).filter(User.id == uid).first()
        return uc[uid]

    buf = BytesIO()
    c   = rl_canvas.Canvas(buf, pagesize=A4)
    page_num = [1]

    def new_page():
        c.showPage()
        page_num[0] += 1

    def draw_footer():
        c.setFont("DV", 7.5)
        c.setFillColor(GREY)
        c.drawString(ML, 1.1 * cm, f"MB Bank — Hệ thống Quản lý Thực tập  |  {ref_no}")
        c.drawRightString(W - MR, 1.1 * cm, f"Trang {page_num[0]}")
        c.setStrokeColor(LGREY)
        c.setLineWidth(0.5)
        c.line(ML, 1.4 * cm, W - MR, 1.4 * cm)

    def check_space(y, need):
        """Return new y (possibly after new page) if not enough space."""
        if y - need < 2 * cm:
            draw_footer()
            new_page()
            return H - MT
        return y

    # ═══════════════════════════════════════════════════════
    # PAGE 1
    # ═══════════════════════════════════════════════════════
    y = H - MT

    # Banner
    banner_h = 1.6 * cm
    c.setFillColor(RED)
    c.rect(ML, y - banner_h, content_w, banner_h, fill=1, stroke=0)
    c.setFillColor(WHITE)
    c.setFont("DV-B", 13)
    c.drawCentredString(W / 2, y - banner_h + 0.55 * cm, "MB MILITARY BANK  —  Hệ thống Quản lý Thực tập")
    y -= banner_h + 0.3 * cm

    # Title
    c.setFont("DV-B", 16)
    c.setFillColor(RED)
    c.drawCentredString(W / 2, y - 0.6 * cm, "BÁO CÁO TỔNG HỢP TÌNH HÌNH THỰC TẬP SINH")
    y -= 1.0 * cm

    # Subtitle
    c.setFont("DV", 9.5)
    c.setFillColor(DARK)
    c.drawCentredString(W / 2, y - 0.1 * cm,
        f"Đợt: {filter_info}  •  Ngày xuất: {gen_time}  •  Mã BC: {ref_no}")
    y -= 0.5 * cm

    # Red underline
    c.setStrokeColor(RED)
    c.setLineWidth(1.5)
    c.line(ML, y, W - MR, y)
    y -= 0.5 * cm

    # ── Section I: Executive Summary ──
    c.setFont("DV-B", 12)
    c.setFillColor(RED)
    c.drawString(ML, y, "I. TÓM TẮT ĐIỀU HÀNH (EXECUTIVE SUMMARY)")
    y -= 0.5 * cm

    # Summary paragraph — draw line by line
    summary = (
        f"Hệ thống ghi nhận {len(profs_f)} thực tập sinh, tỷ lệ hoàn thành nhiệm vụ đạt "
        f"{comp_rate}% ({approved}/{total_tasks} nhiệm vụ đã duyệt), "
        f"điểm đánh giá trung bình {avg_score}/10."
    )
    c.setFont("DV", 10)
    c.setFillColor(colors.black)
    for line in _wrap_text(c, summary, "DV", 10, content_w):
        c.drawString(ML, y, line)
        y -= 0.45 * cm
    y -= 0.2 * cm

    # KPI table
    kpi_rows = [
        ["CHỈ SỐ", "GIÁ TRỊ", "CHỈ SỐ", "GIÁ TRỊ"],
        ["Tổng TTS", str(len(profs_f)),     "Tỷ lệ HT NV",  f"{comp_rate}%"],
        ["Đã đánh giá", str(len(evals_f)),  "Điểm TB",       f"{avg_score}/10"],
        ["Tổng nhiệm vụ", str(total_tasks), "NV quá hạn",    str(overdue_cnt)],
    ]
    kpi_cw = [content_w * 0.32, content_w * 0.18, content_w * 0.32, content_w * 0.18]
    _draw_table(c, ML, y, kpi_rows, kpi_cw, row_height=20, font_size=9)
    y -= len(kpi_rows) * 20 + 0.4 * cm

    # ── Section II: By Batch ──
    y = check_space(y, 3 * cm)
    c.setFont("DV-B", 12)
    c.setFillColor(RED)
    c.drawString(ML, y, "II. THỐNG KÊ THEO ĐỢT THỰC TẬP")
    y -= 0.5 * cm

    batch_rows = [["Tên đợt", "Thời gian", "TTS", "Tổng NV", "HT", "Tỷ lệ", "Đã ĐG", "Điểm TB", "Trạng thái"]]
    for b in batches_q:
        if batch_id and b.id != batch_id:
            continue
        b_iids  = [p.user_id for p in profiles_q if p.batch_id == b.id]
        b_tasks = [t for t in tasks_f if t.batch_id == b.id]
        b_evs   = [e for e in evals_f  if e.intern_id in b_iids]
        bt = len(b_tasks); ba = sum(1 for t in b_tasks if t.status == "approved")
        br = f"{round(ba/bt*100,1)}%" if bt else "—"
        bavg = str(round(sum(e.total_score for e in b_evs)/len(b_evs),1)) if b_evs else "—"
        period = ""
        if b.start_date and b.end_date:
            period = f"{b.start_date.strftime('%d/%m/%y')}-{b.end_date.strftime('%d/%m/%y')}"
        batch_rows.append([
            b.batch_name, period, len(b_iids), bt, ba, br, len(b_evs), bavg,
            "Đang mở" if b.status == "open" else "Đã đóng",
        ])
    b_cw_total = content_w
    b_cw = [
        b_cw_total * 0.22, b_cw_total * 0.17, b_cw_total * 0.07, b_cw_total * 0.08,
        b_cw_total * 0.06, b_cw_total * 0.09, b_cw_total * 0.08, b_cw_total * 0.09,
        b_cw_total * 0.14,
    ]
    needed = len(batch_rows) * 18 + 0.6 * cm
    y = check_space(y, needed)
    _draw_table(c, ML, y, batch_rows, b_cw, row_height=18, font_size=8.5)
    y -= len(batch_rows) * 18 + 0.5 * cm

    # ── Section III: Task status ──
    y = check_space(y, 3 * cm)
    c.setFont("DV-B", 12)
    c.setFillColor(RED)
    c.drawString(ML, y, "III. PHÂN BỐ TRẠNG THÁI NHIỆM VỤ")
    y -= 0.5 * cm

    st_rows = [["Trạng thái", "Số lượng", "Tỷ lệ"]]
    for k, label in [
        ("new","Chưa nộp"), ("submitted","Đã nộp"),
        ("request_change","Cần sửa"), ("approved","Đã duyệt"), ("overdue","Quá hạn"),
    ]:
        cnt  = sum(1 for t in tasks_f if t.status == k)
        rate = f"{round(cnt/total_tasks*100,1)}%" if total_tasks else "0%"
        st_rows.append([label, str(cnt), rate])

    st_cw = [content_w * 0.55, content_w * 0.225, content_w * 0.225]
    _draw_table(c, ML + content_w * 0.1, y, st_rows,
                [content_w * 0.45, content_w * 0.18, content_w * 0.18],
                row_height=18, font_size=9)
    y -= len(st_rows) * 18 + 0.5 * cm

    # ── Section IV: Evaluations ──
    if evals_f:
        y = check_space(y, 3 * cm)
        c.setFont("DV-B", 12)
        c.setFillColor(RED)
        c.drawString(ML, y, "IV. KẾT QUẢ ĐÁNH GIÁ THỰC TẬP SINH")
        y -= 0.5 * cm

        ev_rows = [["Họ tên TTS", "Thái độ", "Kỷ luật", "Học hỏi", "Kỹ năng", "Kết quả", "Tổng", "Xếp loại"]]
        for ev in sorted(evals_f, key=lambda x: -(x.total_score or 0)):
            iu = gu(ev.intern_id)
            cs = ev.criteria_scores or {}
            ev_rows.append([
                iu.full_name if iu else f"ID:{ev.intern_id}",
                str(cs.get("attitude","—")), str(cs.get("discipline","—")),
                str(cs.get("learning","—")),  str(cs.get("skills","—")),
                str(cs.get("task_result","—")), str(ev.total_score), ev.ranking or "—",
            ])
        ev_cw = [content_w*0.27, content_w*0.1, content_w*0.1,
                 content_w*0.1, content_w*0.1, content_w*0.1, content_w*0.1, content_w*0.13]

        chunk_size = 25
        for chunk_start in range(0, len(ev_rows), chunk_size):
            chunk = ([ev_rows[0]] if chunk_start > 0 else []) + ev_rows[chunk_start:chunk_start + chunk_size]
            needed = len(chunk) * 17
            y = check_space(y, needed + 0.5 * cm)
            _draw_table(c, ML, y, chunk, ev_cw, row_height=17, font_size=8.5)
            y -= len(chunk) * 17 + 0.3 * cm

    # ── Section V: Conclusions ──
    y = check_space(y, 5 * cm)
    c.setStrokeColor(LGREY)
    c.setLineWidth(0.5)
    c.line(ML, y + 0.1 * cm, W - MR, y + 0.1 * cm)
    y -= 0.3 * cm
    c.setFont("DV-B", 12)
    c.setFillColor(RED)
    c.drawString(ML, y, "V. KẾT LUẬN VÀ KÝ DUYỆT")
    y -= 0.5 * cm

    conclusion = (
        f"Tỷ lệ hoàn thành nhiệm vụ đạt {comp_rate}%. "
        f"Có {overdue_cnt} nhiệm vụ quá hạn cần theo dõi. "
        f"Điểm đánh giá trung bình {avg_score}/10. "
        "HR đề xuất tăng cường giám sát các TTS có tỷ lệ hoàn thành thấp."
    )
    c.setFont("DV", 10)
    c.setFillColor(colors.black)
    for line in _wrap_text(c, conclusion, "DV", 10, content_w):
        c.drawString(ML, y, line)
        y -= 0.45 * cm
    y -= 0.4 * cm

    # Signature block
    sig_w = (content_w - 0.8 * cm) / 3
    for i, label in enumerate(["Người lập báo cáo", "Trưởng phòng HR", "Ban Quản lý"]):
        sx = ML + i * (sig_w + 0.4 * cm)
        c.setFillColor(LIGHT)
        c.rect(sx, y - 2.2 * cm, sig_w, 2.2 * cm, fill=1, stroke=0)
        c.setStrokeColor(LGREY)
        c.setLineWidth(0.5)
        c.rect(sx, y - 2.2 * cm, sig_w, 2.2 * cm, fill=0, stroke=1)
        c.setFont("DV-B", 9)
        c.setFillColor(DARK)
        c.drawCentredString(sx + sig_w / 2, y - 0.4 * cm, label)
        c.setFont("DV", 8.5)
        c.setFillColor(GREY)
        c.drawCentredString(sx + sig_w / 2, y - 2.0 * cm, "Ký tên & ngày:")
    y -= 2.6 * cm

    # Confidential footer line
    y -= 0.3 * cm
    c.setFont("DV", 8)
    c.setFillColor(GREY)
    c.drawCentredString(W / 2, y,
        f"CONFIDENTIAL — For Management Use Only  |  Ref: {ref_no}  |  Xuất: {gen_time}")

    draw_footer()
    c.save()
    buf.seek(0)

    fname = f"bao-cao-thuc-tap-{datetime.now().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={fname}"})